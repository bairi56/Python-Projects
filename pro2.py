import openpyxl
wb = openpyxl.load_workbook('facebook.xlsx')
ws=wb['Sheet1']
class facebook:
    user_number=[]
    user_passward=[]
    member=[]
    basic_info=[]
    sender_name=[]
    nick=[]
    bio=[]
    work=[]
    education_mem=[]
    friends_request_accepted=[]
    user_message=[]
    user_reply=[]
    search_member=[]
    type_message=[]
    reply_message=[]
    save_post=[]
    save_comments=[]
    post_for_matching=[]
    match_word=[]
    def __init__(self,name,last_name,number,passward,date_of_birth,gender):
        ask_for_signup=int(input("Would you want to sign up"))
        if ask_for_signup==1:
            print("Your account is Successfully registered")
        self.name=name
        self.last_name=last_name
        facebook.member.append(self.name)
        print("Sign up members are",facebook.member)
        wb['Sheet1'].cell(row=1, column=2).value = "MEMBERS"
        for i in range(len(facebook.member)):
            wb['Sheet1'].cell(row=i+2, column=2).value = facebook.member[i]
        wb.save('facebook.xlsx')
        # print(facebook.member)
        self.number = number
        self.passward=passward
        self.date_of_birth=date_of_birth
        self.gender=gender
        # self.email=email
        self.login=False
        search_frnd=False
    def log_in(self):
        enter_number=str(input("Enter your number"))
        enter_passward=str(input("ENter your passward for log_in"))
        if enter_number==self.number and enter_passward==self.passward:
            facebook.user_number.append(enter_number)
            facebook.user_passward.append(enter_passward)
            # print("THE USER NUMBER IS",facebook.user_number)
            # print("Thse user passward is",facebook.user_passward)
            print("You Are Successfully Logged in")
            # self.info=FB.profile_info(facebook)
            self.login=True
            # self.info = FB.profile_info(facebook)
        else:
            print("Enter correct number or passward for loging in")
    def log_out(self):
        if self.login==True:
            out=int(input("would you want to log out"))
            if out==1:
                print("you are log out")
            else:
                print("plz dial correction")
class USER:
    def profile_info(self):
        nick_name = str(input("enter your nick name"))
        write_bio = str(input("enter your profile bio"))
        work_experience=str(input("enter your work experience"))
        education=str(input("enter your work education"))
        print("Your Nickname is...:", nick_name)
        print("Your BIO is...:", write_bio)
        print("Your work experince is...:",work_experience)
        print("Your education is...:",education)
        facebook.nick.append(nick_name)
        facebook.bio.append(write_bio)
        facebook.work.append(work_experience)
        facebook.education_mem.append(education)
        wb['Sheet1'].cell(row=1, column=6).value = "WORK Experiece"
        for i in range(len(facebook.work)):
            wb['Sheet1'].cell(row=i + 2, column=6).value = facebook.work[i]
        wb.save('facebook.xlsx')
        wb['Sheet1'].cell(row=1, column=7).value = "NIck Nme"
        for i in range(len(facebook.nick)):
            wb['Sheet1'].cell(row=i + 2, column=7).value=facebook.nick[i]
        wb.save('facebook.xlsx')
        wb['Sheet1'].cell(row=1, column=8).value = "BIO"
        for i in range(len(facebook.bio)):
            wb['Sheet1'].cell(row=i + 2, column=8).value =facebook.bio[i]
        wb.save('facebook.xlsx')
        wb['Sheet1'].cell(row=1, column=9).value = "EDUCTAION"
        for i in range(len(facebook.education_mem)):
            wb['Sheet1'].cell(row=i + 2, column=9).value =facebook.education_mem[i]
        wb.save('facebook.xlsx')
    def search_friends(self):
        search=str(input("enter a name that you wanted to search::"))
        if search in facebook.member:
            print("The name you search is",search)
            facebook.search_member.append(search)
            wb['Sheet1'].cell(row=1, column=3).value = "SEARCH"
            for i in range(len(facebook.search_member)):
                wb['Sheet1'].cell(row=i + 2, column=3).value = facebook.search_member[i]
            wb.save('facebook.xlsx')
            self.search_frnd=True
        else:
            print("Name you searched is not in friend list..")
    def requests(self):
        request_sender=str(input("Enter a name for sending a friend request..::"))
        sender_passward=str(input("Enter a passward fro sending request...::"))
        confirm_passward=str(input("Enter passward for verify for sending request...::"))
        if sender_passward==confirm_passward:
            print("You can send request")
            ask_for_request=int(input("Enter 0 for sending a request"))
            if ask_for_request==0:
                facebook.sender_name.append(request_sender)
                print("Your request is sended",facebook.sender_name)
                ask_for_request = int(input("Enter 1 for ACCEPTING REQUEST"))
                if ask_for_request==1:
                    facebook.search_member.append(request_sender)
                    facebook.friends_request_accepted.append(request_sender)
                    print("Your request is accepted",facebook.friends_request_accepted)
                    wb['Sheet1'].cell(row=1, column=4).value = "ACCEPTED REQUESTS"
                    for i in range(len(facebook.search_member)):
                        wb['Sheet1'].cell(row=i + 2, column=4).value = facebook.search_member[i]
                    wb.save('facebook.xlsx')
                    ask_for_request = int(input("enter 2 for rejecTING REQUEST"))
                    if ask_for_request==2:
                        facebook.sender_name.remove(request_sender)
                        print("your request is rejected",facebook.sender_name)
        else:
            print("Need correction")
    def message(self):
        name_select = str(input("enter a name to whome you want to send a message"))
        # print("You Are Sending Message to :", name_select)
        if name_select in facebook.search_member:
            print("You Are Sending Message to :", name_select)
            enter_messagge = str(input('TYPE A MESSAGE............:'))
            print("Typed message is",enter_messagge)
            facebook.type_message.append(enter_messagge)
            name_select_reply = str(input("Send Reply.............:"))
            print("Reply message",name_select_reply)
            facebook.reply_message.append(name_select_reply)
            wb['Sheet1'].cell(row=1, column=10).value = "Type_Message"
            for i in range(len(facebook.type_message)):
                wb['Sheet1'].cell(row=i + 2, column=10).value = facebook.type_message[i]
            wb.save('facebook.xlsx')
            wb['Sheet1'].cell(row=1, column=11).value = "ReplY_MessaGe"
            for i in range(len(facebook.reply_message)):
                wb['Sheet1'].cell(row=i + 2, column=11).value =facebook.reply_message[i]
            wb.save('facebook.xlsx')
        else:
            print("This is not in member list")
    def create_post(self):
        ask_first=int(input('Would you want to create a post..?'))
        if ask_first==1:
            print("You can create a post...")
            post_title=str(input("Write your post.................................:"))
            print("The post your created is:",post_title)
            facebook.save_post.append(post_title)
            wb['Sheet1'].cell(row=1, column=12).value = "Save Post"
            for i in range(len(facebook.save_post)):
                wb['Sheet1'].cell(row=i + 2, column=12).value =facebook.save_post[i]
            wb.save('facebook.xlsx')
            self.cmmnt=INSTRUCTIONS.add_comments(USER)
            privacy_share=int(input("TO WHOME Would You Want To Share:==:ENTER 1 if you want to share your post to friends::::: 2 for setting privacy as for only me ::::: 3 for making your post visble to public: "))
            if privacy_share==1:
                print("Your post will be visible for friends only..")
            elif privacy_share==2:
                print("Your post will be visible for only user..")
            elif privacy_share==3:
                print("Your post will be visible to public..")
    def search_post(self):
        match=str(input("Enter a word you want to match"))
        facebook.match_word.append(match)
        wb['Sheet1'].cell(row=1, column=13).value = "WORD Searched"
        for i in range(len(facebook.match_word)):
            wb['Sheet1'].cell(row=i + 2, column=13).value = facebook.match_word[i]
        wb.save('facebook.xlsx')
        matching=[s for s in facebook.save_post if any(xs in s for xs in match)]
        print(matching)
        facebook.post_for_matching.append(matching)
        print(facebook.post_for_matching)

        # wb['Sheet1'].cell(row=1, column=14).value = "Post_For_Matching"
        # for i in range(len(facebook.post_for_matching)):
        #     wb['Sheet1'].cell(row=i + 2, column=14).value = facebook.post_for_matching[i]
        # wb.save('facebook.xlsx')
        # search_choice=str(input("Enter you choice to search posts"))
        # for i in range(len(facebook.save_post)):
        #     if search_choice in facebook.save_post:
        #         print(facebook.save_post)
        #     elif search_choice not in facebook.save_post:
        #         print("PLZ Search correctly")
class INSTRUCTIONS:
    def add_comments(self):
        write_comments=str(input("Write a comment on post"))
        print(write_comments)
        facebook.save_comments.append(write_comments)
        wb['Sheet1'].cell(row=1, column=15).value = "Comment ON Post"
        for i in range(len(facebook.save_comments)):
            wb['Sheet1'].cell(row=i + 2, column=15).value = facebook.save_comments[i]
        wb.save('facebook.xlsx')
    def notification(self):
        #self.frnd=FB.search_friends(posts)
        # if USER.search_frnd==True:
        self.frnd=USER.search_friends(INSTRUCTIONS)
        print("Your searched friend")
        USER.search_frnd=True
        self.info=USER.profile_info(INSTRUCTIONS)
        # self.info = FB.profile_info(posts)
        # self.obj = FB.requests(posts)
        self.obj=USER.requests(INSTRUCTIONS)
        self.info=USER.profile_info(INSTRUCTIONS)
        noti_ask = int(input('Enter 1 for sending a request notification::'))
        if noti_ask==1:
            print("You have a new friend request")
        else:
            print("click true for notifications")
            # self.info=FB.profile_info(posts)
        # self.msg=FB.message(posts)
        self.msg=USER.message(INSTRUCTIONS)
        noti_ask=int(input('Enter 1 for sending a message notification::'))
        if noti_ask==1:
            print("You have a new message")
        else:
            print("click true for notifications")
        # self.comments=FB.create_post(posts)
        self.comments=USER.create_post(INSTRUCTIONS)
        noti_ask=int(input("Enter 1 For Comments Notifications"))
        if noti_ask==1:
            print("Someone commented on your post")
        else:
            print("click true for notifications")
st1=facebook(str(input("Enter your first name")),str(input("Enter your 2nd name")),str(input("Enter your number")),str(input("Enter your Passward")),str(input("Write your Date of birth")),str(input("Select whether you are Male or Female")))
# st1=facebook("ZUBAIR NIAZI","ZUBI","03075270814","ZUBI786","23-12-2003","MALE")
st2=facebook("DANIYAL NIAZI","DANI","03017263359","ZUBI","23-12-2003","MALE")
st3=facebook("SAJID NIAZI","SAJU","03076570814","786","23-12-2003","MALE")
st4=facebook("HUMAIR NIAZI","HUMI","03067816802","Z86","23-12-2003","MALE")
st5=facebook("RIZWAN NIAZI","RIZVI","03194171385","ZU6","23-12-2003","MALE")
st1.log_in()
# st1.log_out()
S=USER()
# S.requests()
# S.requests()
# S.requests()
A=INSTRUCTIONS()
A.notification()
# A.notifications()
S.search_post()
